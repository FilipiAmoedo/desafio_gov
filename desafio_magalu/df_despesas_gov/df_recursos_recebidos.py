import pandas as pd
from openpyxl import load_workbook
from desafio_magalu.api.api_gov import api_gov_despesas
import os


def read_and_save(df):
    wb = load_workbook('dados_gov.xlsx')
    sheet_name = wb.sheetnames[0]
    with pd.ExcelWriter('dados_gov.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        startrow = writer.sheets[sheet_name].max_row
        df.to_excel(writer, startrow=startrow, index=False, header=False)


def save_excel():
    df = pd.DataFrame(api_gov_despesas())
    df.loc[df['siglaUFPessoa'] == '-1', 'siglaUFPessoa'] = 'Sem Informação'
    if os.path.exists('dados_gov.xlsx'):
        read_and_save(df)
    else:
        df.to_excel('dados_gov.xlsx', index=False)


if __name__ == '__main__':
    save_excel()